"""선정 결과를 엑셀로 내보낸다: 선정 27개 / 보종별 요약 / 대체 후보 / 전체 분류."""
from __future__ import annotations
import collections
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from select27 import load, select

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
LIFE_FILL = PatternFill("solid", fgColor="E8EEF7")
NONLIFE_FILL = PatternFill("solid", fgColor="FDF0E6")
NOTE_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncol, row=1):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = f"A{row + 1}"  # ws.cell() 로 지정하면 빈 셀이 만들어져 행이 밀린다


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def body(ws, r, ncol, fill=None):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = Font(name=FONT, size=10)
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(c in (3, ncol)))
        if fill:
            cell.fill = fill


def build(out: Path):
    rows = load()
    chosen, reasons, alts = select(rows)

    wb = Workbook()

    # 1) 선정 27개
    ws = wb.active
    ws.title = "선정 27개"
    head = ["연번", "보험구분", "보종", "판매사명", "상품명", "간편심사", "계약건수", "월납실적(원)", "실적순위", "선정근거"]
    ws.append(head)
    style_header(ws, len(head))
    for i, r in enumerate(chosen, 1):
        ws.append([i, r["sector"], r["line"], r["insurer"], r["name"], "O" if r["simplified"] else "",
                   r["count"], r["premium"], r["rank"], reasons[r["no"]]])
        body(ws, ws.max_row, len(head), LIFE_FILL if r["sector"] == "생명보험" else NONLIFE_FILL)
        ws.cell(row=ws.max_row, column=8).number_format = "#,##0"
    total = ws.max_row + 1
    ws.cell(row=total, column=5, value="합계")
    ws.cell(row=total, column=7, value=f"=SUM(G2:G{total - 1})")
    ws.cell(row=total, column=8, value=f"=SUM(H2:H{total - 1})").number_format = "#,##0"
    for c in (5, 7, 8):
        ws.cell(row=total, column=c).font = Font(name=FONT, size=10, bold=True)
    ws.cell(row=total, column=10, value="합계는 파일을 열 때 계산됩니다(수식 유지).").font = Font(name=FONT, size=9, italic=True)
    autosize(ws, [6, 10, 20, 18, 46, 9, 9, 14, 9, 42])
    note = total + 2
    ws.cell(row=note, column=1, value="선정 기준")
    ws.cell(row=note, column=1).font = Font(name=FONT, size=10, bold=True)
    for k, text in enumerate([
        "1) 월납실적 내림차순(동률 시 계약건수)을 1차 지표로 사용.",
        "2) 생명보험·손해보험 각 보종에서 실적 1위 상품을 먼저 확보하여 모든 보종을 최소 1개씩 포함.",
        "3) 남은 자리는 실적순으로 채우되 같은 판매사×보종 조합은 1개만, 한 보종은 최대 3개까지(종신·연금 편중 방지).",
        "4) 보종은 상품명 키워드로 분류(beluga/classify.py). 원본: iFA 판매실적 상품목록 199건.",
    ], start=1):
        c = ws.cell(row=note + k, column=1, value=text)
        c.font = Font(name=FONT, size=9)
        c.fill = NOTE_FILL

    # 2) 보종별 요약
    ws2 = wb.create_sheet("보종별 요약")
    head2 = ["보험구분", "보종", "전체 상품수", "선정 수", "선정 상품(판매사)"]
    ws2.append(head2)
    style_header(ws2, len(head2))
    all_cnt = collections.Counter((r["sector"], r["line"]) for r in rows)
    for (sec, ln), n in sorted(all_cnt.items()):
        picks = [c for c in chosen if (c["sector"], c["line"]) == (sec, ln)]
        ws2.append([sec, ln, n, len(picks), " / ".join(f"{p['name']}({p['insurer']})" for p in picks)])
        body(ws2, ws2.max_row, len(head2), LIFE_FILL if sec == "생명보험" else NONLIFE_FILL)
    autosize(ws2, [10, 22, 12, 9, 80])

    # 3) 대체 후보
    ws3 = wb.create_sheet("대체 후보")
    head3 = ["선정 상품", "보험구분", "보종", "대체 후보 1", "대체 후보 2"]
    ws3.append(head3)
    style_header(ws3, len(head3))
    for r in chosen:
        a = alts[r["no"]]
        ws3.append([r["name"], r["sector"], r["line"],
                    f"{a[0]['name']}({a[0]['insurer']})" if len(a) > 0 else "-",
                    f"{a[1]['name']}({a[1]['insurer']})" if len(a) > 1 else "-"])
        body(ws3, ws3.max_row, len(head3), LIFE_FILL if r["sector"] == "생명보험" else NONLIFE_FILL)
    autosize(ws3, [42, 10, 20, 42, 42])

    # 4) 전체 분류 (검산용)
    ws4 = wb.create_sheet("전체 분류")
    head4 = ["원번호", "실적순위", "보험구분", "보종", "판매사명", "상품명", "간편심사", "계약건수", "월납실적(원)", "선정"]
    ws4.append(head4)
    style_header(ws4, len(head4))
    picked = {r["no"] for r in chosen}
    for r in rows:
        ws4.append([r["no"], r["rank"], r["sector"], r["line"], r["insurer"], r["name"],
                    "O" if r["simplified"] else "", r["count"], r["premium"], "★" if r["no"] in picked else ""])
        body(ws4, ws4.max_row, len(head4))
        ws4.cell(row=ws4.max_row, column=9).number_format = "#,##0"
    ws4.auto_filter.ref = f"A1:J{ws4.max_row}"
    autosize(ws4, [8, 9, 10, 22, 18, 52, 9, 9, 14, 6])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out, chosen


if __name__ == "__main__":
    out, chosen = build(Path("out/벨루가_보종별_상품27선.xlsx"))
    print(f"{out}  ({len(chosen)}개)")
