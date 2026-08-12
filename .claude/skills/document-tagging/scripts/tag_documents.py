#!/usr/bin/env python3
"""문서 제목(파일명) 기준으로 표준 태그를 부여한다.

사용법:
    python3 tag_documents.py 입력.xlsx -o 결과.xlsx
    python3 tag_documents.py 입력.csv  -o 결과.csv --rules tag_rules.json

핵심 설계:
  * 판단 기준은 항상 "제목"이다. 본문을 열지 않고도 재현 가능한 태깅을 만든다.
  * 태그는 tag_rules.json에 등록된 표준 태그만 사용한다. 규칙에 없는 개념이
    반복적으로 등장하면 규칙 파일에 태그를 "추가"하는 방식으로만 확장한다.
    이렇게 해야 문서마다 새 태그가 생겨나는 일을 구조적으로 막을 수 있다.
  * 태그는 최대 3개. 우선순위(priority)가 낮은 규칙이 먼저 선택된다.
"""

import argparse
import json
import os
import re
import sys
from collections import Counter

MAX_TAGS = 3

# 등록 시 자동으로 붙는 접미사(_260812_S337_171152, _251112_09815594 등)와 확장자를 제거한다.
SUFFIX_PATTERNS = [
    r"_\d{6}_[A-Za-z가-힣]+[^_]*_\d{6}$",   # _260812_S337_171152 / _260623_CX_120546
    r"_\d{6}_\d{6,}$",
    r"_\d{8}$",
    r"_\d{7,}$",                             # _05264888 (문서 일련번호)
]


def normalize_title(name: str) -> str:
    """파일명에서 확장자와 등록 자동 접미사를 떼어낸 '판단용 제목'을 만든다."""
    title = (name or "").strip()
    title = re.sub(r"\.[A-Za-z0-9]{1,5}$", "", title)
    for _ in range(3):
        before = title
        for pat in SUFFIX_PATTERNS:
            title = re.sub(pat, "", title)
        if title == before:
            break
    return title.strip(" _-")


def load_rules(path: str) -> dict:
    with open(path, encoding="utf-8") as fh:
        rules = json.load(fh)
    for rule in rules["tags"]:
        rule["_re"] = [re.compile(p, re.IGNORECASE) for p in rule["patterns"]]
        rule["_not"] = [re.compile(p, re.IGNORECASE) for p in rule.get("not", [])]
    for rule in rules["exclude"]:
        rule["_re"] = [re.compile(p, re.IGNORECASE) for p in rule["patterns"]]
        rule["_except"] = [re.compile(p, re.IGNORECASE) for p in rule.get("except", [])]
    for rule in rules["fallback"]:
        rule["_re"] = re.compile(rule["match"])
    bh = rules.get("book_heuristic", {})
    bh["_keep"] = [re.compile(p) for p in bh.get("keep_patterns", [])]
    return rules


def check_excluded(title: str, rules: dict):
    """제외 대상이면 (True, 사유)를 돌려준다."""
    for rule in rules["exclude"]:
        if any(r.search(title) for r in rule["_re"]):
            if any(e.search(title) for e in rule["_except"]):
                continue
            return True, rule["reason"]

    bh = rules.get("book_heuristic", {})
    if bh.get("apply_when_no_separator") and title and "_" not in title:
        if not any(k.search(title) for k in bh["_keep"]):
            return True, "도서"
    return False, ""


def pick_tags(title: str, rules: dict):
    """제목에서 매칭된 표준 태그를 우선순위 순으로 최대 3개 고른다."""
    hits = []
    for rule in rules["tags"]:
        if any(n.search(title) for n in rule["_not"]):
            continue
        for rx in rule["_re"]:
            m = rx.search(title)
            if m:
                hits.append((rule["priority"], m.start(), rule["tag"], m.group(0).strip()))
                break

    hits.sort(key=lambda h: (h[0], h[1]))
    tags, evidence = [], []
    for _, _, tag, keyword in hits:
        if tag in tags:
            continue
        tags.append(tag)
        evidence.append(keyword)
        if len(tags) == MAX_TAGS:
            break

    if tags:
        return tags, "제목의 '" + "', '".join(evidence) + "' 기준"

    for rule in rules["fallback"]:
        if rule["_re"].search(title):
            return list(rule["tags"]), "제목의 분류 체계(%s) 기준" % rule["match"].strip("^$.")
    return [], "판단 근거 없음"


def tag_one(name: str, rules: dict) -> dict:
    title = normalize_title(name)
    excluded, reason = check_excluded(title, rules)
    if excluded:
        return {
            "제목": title,
            "태깅여부": "제외",
            "태그1": "", "태그2": "", "태그3": "",
            "판단근거": "%s에 해당하여 태깅 제외" % reason,
        }
    tags, why = pick_tags(title, rules)
    tags += [""] * (MAX_TAGS - len(tags))
    return {
        "제목": title,
        "태깅여부": "대상",
        "태그1": tags[0], "태그2": tags[1], "태그3": tags[2],
        "판단근거": why,
    }


# --------------------------------------------------------------------------- I/O

def read_rows(path: str):
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        ws = openpyxl.load_workbook(path, read_only=True, data_only=True).worksheets[0]
        it = ws.iter_rows(values_only=True)
        header = ["" if c is None else str(c) for c in next(it)]
        return header, [["" if c is None else str(c) for c in row] for row in it]

    import csv
    with open(path, encoding="utf-8-sig", newline="") as fh:
        reader = list(csv.reader(fh))
    return reader[0], reader[1:]


def find_name_columns(header):
    """제목으로 쓸 컬럼(우선순위 순)을 찾는다. 없으면 첫 컬럼을 쓴다."""
    candidates = ["파일명", "게시물제목", "제목", "문서명", "filename", "file", "title", "name"]
    idx = []
    for want in candidates:
        for i, col in enumerate(header):
            if col.strip().lower() == want.lower() and i not in idx:
                idx.append(i)
    return idx or [0]


def write_rows(path, header, rows):
    if path.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "태깅결과"
        ws.append(header)
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        for i, col in enumerate(header, start=1):
            width = 60 if col in ("파일명", "제목") else max(10, min(28, len(col) * 2 + 6))
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        wb.save(path)
    else:
        import csv
        with open(path, "w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(header)
            writer.writerows(rows)


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser(description="문서 제목 기준 표준 태그 부여")
    ap.add_argument("input", help="문서 목록 파일 (.xlsx 또는 .csv)")
    ap.add_argument("-o", "--output", required=True, help="결과 파일 (.xlsx 또는 .csv)")
    ap.add_argument("--rules", default=os.path.join(here, "tag_rules.json"))
    ap.add_argument("--name-column", help="제목으로 사용할 컬럼명 (미지정 시 자동 탐색)")
    ap.add_argument("--json-out", help="검토 스크립트용 태깅 결과 JSON 경로")
    ap.add_argument("--text-out", help="문서별 출력 형식(문서명/태깅여부/태그1~3/판단근거) 텍스트 경로")
    args = ap.parse_args()

    rules = load_rules(args.rules)
    header, rows = read_rows(args.input)

    if args.name_column:
        if args.name_column not in header:
            sys.exit("컬럼을 찾을 수 없습니다: %s (있는 컬럼: %s)" % (args.name_column, ", ".join(header)))
        name_idx = [header.index(args.name_column)]
    else:
        name_idx = find_name_columns(header)

    added = ["태깅여부", "태그1", "태그2", "태그3", "판단근거"]
    out_header = list(header) + added
    out_rows, records = [], []
    for row in rows:
        row = list(row) + [""] * (len(header) - len(row))
        name = next((row[i] for i in name_idx if i < len(row) and row[i].strip()), "")
        result = tag_one(name, rules)
        out_rows.append(row + [result[k] for k in added])
        records.append({
            "제목": result["제목"],
            "원본파일명": name,
            "태깅여부": result["태깅여부"],
            "태그": [t for t in (result["태그1"], result["태그2"], result["태그3"]) if t],
            "판단근거": result["판단근거"],
        })

    write_rows(args.output, out_header, out_rows)

    if args.text_out:
        with open(args.text_out, "w", encoding="utf-8") as fh:
            for rec in records:
                tags = rec["태그"] + [""] * (MAX_TAGS - len(rec["태그"]))
                fh.write("문서명: %s\n\n" % (rec["원본파일명"] or rec["제목"]))
                fh.write("태깅여부: %s\n\n" % rec["태깅여부"])
                for i, t in enumerate(tags, start=1):
                    fh.write("태그%d: %s\n" % (i, t))
                fh.write("\n간단한 판단근거: %s\n\n%s\n\n" % (rec["판단근거"], "-" * 60))

    if args.json_out:
        with open(args.json_out, "w", encoding="utf-8") as fh:
            json.dump(records, fh, ensure_ascii=False)

    targets = [r for r in records if r["태깅여부"] == "대상"]
    counter = Counter(t for r in targets for t in r["태그"])
    print("전체 %d건 / 태깅대상 %d건 / 제외 %d건" % (len(records), len(targets), len(records) - len(targets)))
    print("사용된 태그 종류: %d개" % len(counter))
    print("평균 태그 수: %.2f개" % (sum(len(r["태그"]) for r in targets) / max(1, len(targets))))
    print("상위 태그: " + ", ".join("#%s(%d)" % (k, v) for k, v in counter.most_common(15)))
    print("결과 저장: %s" % args.output)


if __name__ == "__main__":
    main()
